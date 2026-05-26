import numpy as np
from utils.excel_utils import _autosize
from openpyxl.chart import PieChart, BarChart, LineChart, Reference
from openpyxl.chart.axis import ChartLines
from openpyxl.styles import Font
from openpyxl.utils.dataframe import dataframe_to_rows

def build_zscore_pivot(z_df):
    """
    Pivot สำหรับ Dashboard:
    บริษัท | ปี1 | ปี2 | ปี3 ... | เฉลี่ย | แนวโน้ม
    """
    pv = (
        z_df
        .pivot_table(
            index="บริษัท",
            columns="ปี",
            values="Z'-Score",
            aggfunc="last"
        )
        .sort_index()
    )

    
    year_cols = [
        c for c in pv.columns
        if isinstance(c, (int, float, np.integer))
    ]
    year_cols = sorted(year_cols)

    # ===== ค่าเฉลี่ย =====
    pv["เฉลี่ย"] = pv[year_cols].mean(axis=1)

    # ===== แนวโน้ม (Linear Trend / Slope) =====
    trends = []

    for _, row in pv.iterrows():
        y = row[year_cols].values.astype(float)
        x = np.arange(len(y))

        # ต้องมีอย่างน้อย 2 ปี
        if np.count_nonzero(~np.isnan(y)) < 2:
            trends.append("N/A")
            continue

        slope = np.polyfit(x[~np.isnan(y)], y[~np.isnan(y)], 1)[0]

        if slope > 0.05:
            trends.append("📈 ดีขึ้น")
        elif slope < -0.05:
            trends.append("📉 แย่ลง")
        else:
            trends.append("➖ ทรงตัว")

    pv["แนวโน้ม"] = trends

    return pv.reset_index()

def add_summary_charts(workbook):
    if "Z_SCORE" not in workbook.sheetnames:
        return

    ws_data = workbook["Z_SCORE"]

    # ลบ SUMMARY เดิมถ้ามี
    if "SUMMARY" in workbook.sheetnames:
        del workbook["SUMMARY"]

    ws = workbook.create_sheet("SUMMARY")

    headers = [c.value for c in ws_data[1]]
    idx = {h: i + 1 for i, h in enumerate(headers)}

    status_col = idx.get("สถานะ")
    year_col = idx.get("ปี")
    zcol = idx.get("Z'-Score")
    company_col = idx.get("บริษัท")

    # -------------------------
    # 1) STATUS SUMMARY
    # -------------------------
    ws["A1"] = "สรุปสถานะบริษัท"
    ws.append(["สถานะ", "จำนวน"])

    status_count = {}
    for r in range(2, ws_data.max_row + 1):
        s = ws_data.cell(r, status_col).value
        if s:
            status_count[s] = status_count.get(s, 0) + 1

    start_row = ws.max_row + 1
    for k, v in status_count.items():
        ws.append([k, v])

    pie = PieChart()
    pie.title = "สัดส่วนสถานะบริษัท"
    data = Reference(ws, min_col=2, min_row=start_row, max_row=ws.max_row)
    labels = Reference(ws, min_col=1, min_row=start_row + 1, max_row=ws.max_row)
    pie.add_data(data, titles_from_data=True)
    pie.set_categories(labels)
    ws.add_chart(pie, "D2")

    # -------------------------
    # 2) AVG Z-SCORE BY YEAR
    # -------------------------
    ws["A10"] = "Z-Score เฉลี่ยรายปี"
    ws.append(["ปี", "Z-Score เฉลี่ย"])

    z_by_year = {}
    cnt_by_year = {}

    for r in range(2, ws_data.max_row + 1):
        y = ws_data.cell(r, year_col).value
        z = ws_data.cell(r, zcol).value
        if y and z is not None:
            z_by_year[y] = z_by_year.get(y, 0) + z
            cnt_by_year[y] = cnt_by_year.get(y, 0) + 1

    start_row = ws.max_row + 1
    for y in sorted(z_by_year):
        ws.append([y, round(z_by_year[y] / cnt_by_year[y], 2)])

    line = LineChart()
    line.title = "Z-Score เฉลี่ยรายปี"
    line.y_axis.title = "Z-Score"
    line.x_axis.title = "ปี"

    data = Reference(ws, min_col=2, min_row=start_row, max_row=ws.max_row)
    cats = Reference(ws, min_col=1, min_row=start_row + 1, max_row=ws.max_row)
    line.add_data(data, titles_from_data=True)
    line.set_categories(cats)

    ws.add_chart(line, "D16")

    
    # -------------------------
    # 3) TREND SUMMARY (ง่าย ๆ)
    # -------------------------
    ws["A20"] = "แนวโน้ม Z-Score (ภาพรวม)"
    ws["A20"].font = Font(bold=True)
    ws.append(["ผลลัพธ์", "เงื่อนไข"])

    # ใช้ปีแรกกับปีล่าสุดเท่านั้น
    years = sorted(z_by_year.keys())
    if len(years) >= 2:
        first_year = years[0]
        last_year = years[-1]

        avg_first = z_by_year[first_year] / cnt_by_year[first_year]
        avg_last  = z_by_year[last_year] / cnt_by_year[last_year]

        diff = avg_last - avg_first

        if diff > 0.05:
            trend = "📈 ดีขึ้น"
        elif diff < -0.05:
            trend = "📉 แย่ลง"
        else:
            trend = "➖ ทรงตัว"

        ws.append([trend, f"{first_year} → {last_year}"])
    else:
        ws.append(["N/A", "ข้อมูลไม่พอ"])

def add_company_zscore_comparison(workbook):
    if "Z_SCORE" not in workbook.sheetnames:
        return

    ws_data = workbook["Z_SCORE"]

    if "DASHBOARD" not in workbook.sheetnames:
        return

    ws = workbook["DASHBOARD"]
    start_row = ws.max_row + 3

    ws[f"A{start_row}"] = "📉 เปรียบเทียบ Z'-Score ล่าสุด (เรียงจากเสี่ยง → ปลอดภัย)"
    ws[f"A{start_row}"].font = Font(bold=True)

    header_row = start_row + 1
    ws.append(["บริษัท", "Z-Score"])

    headers = [c.value for c in ws_data[1]]
    idx = {h: i + 1 for i, h in enumerate(headers)}

    company_col = idx.get("บริษัท")
    year_col = idx.get("ปี")
    z_col = idx.get("Z'-Score")

    if not all([company_col, year_col, z_col]):
        return

    # -----------------------------
    # เลือก Z‑Score ล่าสุดต่อบริษัท
    # -----------------------------
    latest = {}

    for r in range(2, ws_data.max_row + 1):
        cname = ws_data.cell(r, company_col).value
        year = ws_data.cell(r, year_col).value
        z = ws_data.cell(r, z_col).value

        if cname and year and z is not None:
            if cname not in latest or year > latest[cname]["year"]:
                latest[cname] = {"year": year, "z": float(z)}

    if not latest:
        return

    # -----------------------------
    # เรียงตาม Z‑Score (ต่ำ → สูง)
    # -----------------------------
    items = sorted(latest.items(), key=lambda x: x[1]["z"])

    data_start_row = ws.max_row + 1
    for cname, v in items:
        z = max(min(v["z"], 6), -2)
        ws.append([cname, round(z, 2)])

    data_end_row = ws.max_row

    if data_end_row <= data_start_row:
        return

    # -----------------------------
    # Bar chart
    # -----------------------------
    chart = BarChart()
    chart.type = "col"
    chart.style = 10
    chart.title = "Z'-Score ล่าสุด (เรียงตามความเสี่ยง)"
    chart.y_axis.title = "Z-Score"
    chart.x_axis.title = "บริษัท"
    chart.width = 26
    chart.height = 14

    data = Reference(
        ws,
        min_col=2,
        min_row=header_row,
        max_row=data_end_row
    )
    cats = Reference(
        ws,
        min_col=1,
        min_row=data_start_row,
        max_row=data_end_row
    )

    chart.add_data(data, titles_from_data=True)
    chart.set_categories(cats)
    chart.y_axis.majorGridlines = ChartLines()

    ws.add_chart(chart, f"D{start_row}")
    
def add_dashboard_sheet(workbook, pivot_df):
    """
    Dashboard แบบอ่านง่าย:
    - ตารางค่าเฉลี่ย + แนวโน้ม ต่อบริษัท
    """
    if pivot_df is None or pivot_df.empty:
        return

    if "DASHBOARD" in workbook.sheetnames:
        del workbook["DASHBOARD"]

    ws = workbook.create_sheet("DASHBOARD")

    ws["A1"] = "📊 Dashboard: ค่าเฉลี่ย & แนวโน้ม Z'-Score"
    ws["A1"].font = Font(size=13, bold=True)

    ws["A3"] = "บริษัท | Z‑Score เฉลี่ย | แนวโน้ม"
    ws["A3"].font = Font(bold=True)

    # ใช้เฉพาะคอลัมน์ที่ต้องการ
    cols = ["บริษัท", "เฉลี่ย", "แนวโน้ม"]
    view_df = pivot_df[cols].copy()

    for r in dataframe_to_rows(view_df, index=False, header=True):
        ws.append(r)

    _autosize(ws)


def add_company_trend_chart(ws, start_row, pivot_df):

    year_cols = [
        i + 1 for i, c in enumerate(pivot_df.columns)
        if isinstance(c, (int, float))
    ]

    if not year_cols:
        return

    chart = LineChart()
    chart.title = "เปรียบเทียบ Z'-Score รายบริษัท"
    chart.y_axis.title = "Z-Score"
    chart.x_axis.title = "ปี"

    data = Reference(
        ws,
        min_col=year_cols[0],
        min_row=start_row + 1,
        max_col=year_cols[-1],
        max_row=start_row + 1   # แถวบริษัทเดียว
    )

    cats = Reference(
        ws,
        min_col=year_cols[0],
        min_row=start_row,
        max_col=year_cols[-1],
        max_row=start_row
    )

    chart.add_data(data, from_rows=True)
    chart.set_categories(cats)

    ws.add_chart(chart, "J9")