from datetime import datetime
from openpyxl.utils.dataframe import dataframe_to_rows
import os
import pandas as pd
from openpyxl import Workbook, load_workbook

from reports.dashboard import (
    add_summary_charts,
    add_dashboard_sheet,
    add_company_zscore_comparison,
    build_zscore_pivot,
)

from zscore.formatter import _finalize_zdf
from utils.excel_utils import _autosize, _format_numbers, _colorize
from utils.logging_utils import log
from config import Z_SCORE_FILE, Z_SCORE_SHEET, OUTPUT_DIR


def save_zscore_single_sheet(new_df):
    if new_df is None or new_df.empty:
        log("⚠️ Z-Score empty → skip save")
        return

    # ---------- Load / merge workbook ----------
    if os.path.exists(Z_SCORE_FILE):
        try:
            wb = load_workbook(Z_SCORE_FILE)
        except PermissionError:
            alt = os.path.join(
                OUTPUT_DIR,
                f"zscore_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
            )
            log(f"⚠️ zscore.xlsx locked → save {os.path.basename(alt)}")
            _save_new_workbook(new_df, alt)
            return

        if Z_SCORE_SHEET in wb.sheetnames:
            old_df = pd.read_excel(Z_SCORE_FILE, sheet_name=Z_SCORE_SHEET)
            combined = pd.concat([old_df, new_df], ignore_index=True)
        else:
            combined = new_df.copy()
    else:
        wb = Workbook()
        if "Sheet" in wb.sheetnames:
            del wb["Sheet"]
        combined = new_df.copy()

    # ---------- Finalize + dedup ----------
    combined = _finalize_zdf(combined)
    combined = combined.drop_duplicates(
        ["บริษัท", "ปี"],
        keep="last"
    ).reset_index(drop=True)

    # ---------- Recreate Z_SCORE sheet ----------
    if Z_SCORE_SHEET in wb.sheetnames:
        del wb[Z_SCORE_SHEET]

    ws = wb.create_sheet(Z_SCORE_SHEET, 0)
    for r in dataframe_to_rows(combined, index=False, header=True):
        ws.append(r)

    ws.freeze_panes = "A2"
    _autosize(ws)
    _format_numbers(ws)
    _colorize(ws)

    # ==================================================
    # ✅ IMPORTANT PART (ที่คุณถาม)
    # ==================================================

    # 1) SUMMARY : ค่าเฉลี่ย + แนวโน้มภาพรวม
    add_summary_charts(wb)

    # 2) DASHBOARD : ค่าเฉลี่ย + แนวโน้ม ต่อบริษัท
    pivot_df = build_zscore_pivot(combined)
    add_dashboard_sheet(wb, pivot_df)

    # 3) เปรียบเทียบ Z‑Score ล่าสุด (บริษัทเสี่ยง/ปลอดภัย)
    add_company_zscore_comparison(wb)

    # ==================================================
    # ✅ SAVE WORKBOOK (ต้องอยู่ท้ายสุด)
    # ==================================================
    tmp_path = Z_SCORE_FILE.with_suffix(".xlsx.tmp")
    try:
        wb.save(tmp_path)
        os.replace(tmp_path, Z_SCORE_FILE)
    except PermissionError:
        try:
            if os.path.exists(tmp_path):
                os.remove(tmp_path)
        except:
            pass
        alt = os.path.join(
            OUTPUT_DIR,
            f"zscore_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
        )
        log(f"⚠️ zscore.xlsx locked → save {os.path.basename(alt)}")
        wb.save(alt)

def _save_new_workbook(df, path):
    wb = Workbook()
    if "Sheet" in wb.sheetnames:
        del wb["Sheet"]

    ws = wb.create_sheet("Z_SCORE", 0)
    df = _finalize_zdf(df)

    for r in dataframe_to_rows(df, index=False, header=True):
        ws.append(r)

    ws.freeze_panes = "A2"
    _autosize(ws)
    _format_numbers(ws)
    _colorize(ws)

    wb.save(path)
