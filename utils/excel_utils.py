import time
from pathlib import Path
from openpyxl.styles import PatternFill
from zscore.formatter import MONEY_COLS, Z_COL, RATIO_COLS


def safe_to_excel(df, path: Path, retries=5):
    path.parent.mkdir(parents=True, exist_ok=True)

    for i in range(retries):
        try:
            tmp = path.with_suffix(".tmp.xlsx")
            df.to_excel(tmp, index=False, engine="openpyxl")
            tmp.replace(path)
            return
        except PermissionError:
            time.sleep(0.3)

    raise PermissionError(f"Cannot write Excel file: {path}")

def _autosize(ws):
    for col in ws.columns:
        max_len = 0
        col_letter = col[0].column_letter
        for cell in col:
            v = "" if cell.value is None else str(cell.value)
            if len(v) > max_len:
                max_len = len(v)
        ws.column_dimensions[col_letter].width = min(max_len + 2, 32)

def _format_numbers(ws):
    headers = [c.value for c in ws[1]]
    idx = {h: i + 1 for i, h in enumerate(headers)}

    for r in range(2, ws.max_row + 1):
        for h in MONEY_COLS:
            if h in idx:
                ws.cell(r, idx[h]).number_format = "#,##0.00"
        for h in RATIO_COLS:
            if h in idx:
                ws.cell(r, idx[h]).number_format = "0.0000"
        if Z_COL in idx:
            ws.cell(r, idx[Z_COL]).number_format = "0.00"


def _colorize(ws):
    headers = [c.value for c in ws[1]]
    if "สถานะ" not in headers:
        return
    status_col = headers.index("สถานะ") + 1

    fills = {
        "ปลอดภัย": PatternFill("solid", fgColor="C6EFCE"),
        "เฝ้าระวัง": PatternFill("solid", fgColor="FFEB9C"),
        "อันตราย": PatternFill("solid", fgColor="FFC7CE"),
        "ไม่ทราบ": PatternFill("solid", fgColor="D9D9D9"),
    }

    for r in range(2, ws.max_row + 1):
        status = ws.cell(r, status_col).value
        fill = fills.get(status)
        if fill:
            for c in range(1, ws.max_column + 1):
                ws.cell(r, c).fill = fill