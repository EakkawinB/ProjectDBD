import os
import re
import traceback
from datetime import datetime
from datetime import time as dtime
from datetime import timedelta
import base64
import zipfile
import time
import numpy as np
import pandas as pd     
import requests
from pathlib import Path
from playwright.sync_api import sync_playwright, Error as PlaywrightError
from openpyxl import load_workbook, Workbook
from openpyxl.chart import PieChart, BarChart, LineChart, Reference
from openpyxl.chart.axis import ChartLines
from openpyxl.styles import PatternFill
from openpyxl.styles import Font
from openpyxl.utils.dataframe import dataframe_to_rows

BASE_DIR = Path(__file__).resolve().parent
os.chdir(BASE_DIR)

INPUT_IDS_FILE = "company_ids.xlsx"
IDS_SHEET = "ids"
ID_COL = "company_id"

# ===== Time Window Control =====
RUN_START_TIME = dtime(14, 00)    
RUN_END_TIME   = dtime(21, 30)   

OUTPUT_DIR = BASE_DIR / "output"
OUTPUT_DIR.mkdir(exist_ok=True)

Z_SCORE_FILE = OUTPUT_DIR / "zscore.xlsx"
Z_SCORE_SHEET = "Z_SCORE"

HEADLESS = False                
MAX_RETRY = 2
WATCHDOG_RESTART_EVERY = 40         

COMPANY_HARD_TIMEOUT = 3 * 60    # 3 นาทีต่อบริษัท

TEST_MODE = False                # True = เทส / False = รันจริง
TEST_LIMIT = None                # จำนวนบริษัทที่จะเทส 
BASE_URL = "https://datawarehouse.dbd.go.th/"
SAFE_MODE = True

DEFAULT_TIMEOUT = 30000

from datetime import timedelta

def format_elapsed(start_dt, now_dt):
    elapsed = now_dt - start_dt
    total_sec = int(elapsed.total_seconds())

    days = total_sec // 86400
    hours = (total_sec % 86400) // 3600
    minutes = (total_sec % 3600) // 60

    return f"{days} วัน {hours} ชม {minutes} นาที", elapsed


def estimate_avg_time_per_company(start_dt, now_dt, processed):
    if processed <= 0:
        return None
    return (now_dt - start_dt).total_seconds() / processed


def estimate_eta(now_dt, remaining_seconds):
    cur = now_dt
    remaining = remaining_seconds

    while remaining > 0:
        cur_time = cur.time()

        if cur_time < RUN_START_TIME:
            cur = datetime.combine(cur.date(), RUN_START_TIME)
        elif cur_time > RUN_END_TIME:
            cur = datetime.combine(cur.date() + timedelta(days=1), RUN_START_TIME)

        end_today = datetime.combine(cur.date(), RUN_END_TIME)
        usable_today = (end_today - cur).total_seconds()

        if remaining <= usable_today:
            return cur + timedelta(seconds=remaining)

        remaining -= usable_today
        cur = datetime.combine(cur.date() + timedelta(days=1), RUN_START_TIME)

    return cur

COL_ORDER = [
    "บริษัท", "ปี",
    "Current Assets", "Current Liab", "Total Assets", "Retained Earn",
    "Net Income", "Interest Exp", "Tax", "Total Liab",
    "Sales", "Working Cap", "EBIT", "Equity",
    "X1", "X2", "X3", "X4", "X5", "Z'-Score", "สถานะ"
]

MONEY_COLS = [
    "Current Assets", "Current Liab", "Total Assets", "Retained Earn",
    "Net Income", "Interest Exp", "Tax", "Total Liab",
    "Sales", "Working Cap", "EBIT", "Equity"
]
RATIO_COLS = ["X1", "X2", "X3", "X4", "X5"]
Z_COL = "Z'-Score"

def ts():
    return datetime.now().strftime("%Y-%m-%d %H:%M:%S")

def log(msg):
    print(f"[{ts()}] {msg}", flush=True)


# =========================
# HUMAN LIKE SPEED
# =========================
ACTION_DELAY = 0.15      
CLICK_DELAY = 0.25
TYPE_DELAY = 35         # ms ต่อ key

def slow_action(delay=ACTION_DELAY):
    time.sleep(delay)

def slow_click(locator, delay=CLICK_DELAY, force=True):
    locator.click(force=force)
    time.sleep(delay)

def slow_type(locator, text, delay=TYPE_DELAY):
    locator.type(text, delay=delay)
    time.sleep(ACTION_DELAY)

# =====================================================
# PAGE SAFETY
# =====================================================

from playwright.sync_api import Error as PlaywrightError

def ensure_page_alive(page, browser):
    """
    ✅ ดูแลเฉพาะ page
    ❌ ถ้า browser ตาย → ห้าม revive เอง
    """
    try:
        if page.is_closed():
            raise RuntimeError("page closed")

        page.title()
        return page

    except PlaywrightError as e:
        # ✅ สำคัญมาก: ถ้าเป็น TargetClosedError
        if "TargetClosedError" in str(e):
            log("🔴 Browser is closed → cannot recover page here")
            raise   

        # ✅ page พัง แต่ browser ยังอยู่ → สร้าง page ใหม่
        log("♻ Page closed → recreate page")
        page = browser.new_page()
        page.set_default_timeout(DEFAULT_TIMEOUT)
        page.goto(BASE_URL, wait_until="domcontentloaded")
        time.sleep(1.2)
        return page

    except Exception:
        # fallback
        log("♻ Page error → try recreate page")
        page = browser.new_page()
        page.set_default_timeout(DEFAULT_TIMEOUT)
        page.goto(BASE_URL, wait_until="domcontentloaded")
        time.sleep(1.2)
        return page
    
def ensure_home_search_ready(page, browser, timeout=180):
    """
    ✅ Robust version for Scheduler / Server
    """

    start = time.time()
    last_state = ""

    while time.time() - start < timeout:
        try:
            # ถ้า page ตาย → revive
            page = ensure_page_alive(page, browser)

            # บังคับกลับ Home เสมอ
            if BASE_URL not in page.url:
                page.goto(BASE_URL, wait_until="domcontentloaded")
                page.wait_for_timeout(2500)

            popup_gate(page, soft=True)
            close_popups(page)
            page.wait_for_timeout(500)

            # ✅ selector แบบ flexible (สำคัญ)
            candidates = [
                "input.form-control",
                "input[type='search']",
                "form input",
            ]

            for sel in candidates:
                loc = page.locator(sel)
                if loc.count() > 0:
                    box = loc.first
                    if box.is_visible() and box.is_enabled():
                        log(f"✅ Home search ready (selector: {sel})")
                        return box

            last_state = f"url={page.url}"

        except Exception as e:
            last_state = str(e)

        page.wait_for_timeout(2000)

    raise TimeoutError(f"❌ Home search box not available ({last_state})")

def norm_text(s):
    if s is None:
        return ""
    s = str(s).replace("\u00a0", " ").replace("\u200b", " ")
    return re.sub(r"\s+", " ", s).strip()


def to_float(txt):
    if txt is None:
        return None
    t = norm_text(txt)
    if t in ["", "-", "-"]:
        return None
    neg = False
    if t.startswith("(") and t.endswith(")"):
        neg = True
        t = t[1:-1].strip()
    t = t.replace(",", "")
    m = re.search(r"-?\d+(\.\d+)?", t)
    if not m:
        return None
    try:
        v = float(m.group(0))
        if neg:
            v = -v
        return round(v, 2)
    except:
        return None    
    
def looks_like_person_name(name):
    if not name:
        return False
    name = name.strip()
    prefixes = ["นาย", "นาง", "นางสาว", "Mr.", "Mrs.", "Ms."]
    return any(name.startswith(p) for p in prefixes)

def looks_like_university(name):
    if not name:
        return False

    keywords = [
        "มหาวิทยาลัย",
        "University",
        "สถาบัน",
        "Institute",
        "College",
        "ราชภัฏ",
        "อาชีว",
        "วิทยาลัย"
    ]

    name_lower = name.lower()
    return any(k.lower() in name_lower for k in keywords)

def normalize_company_id(cid):
    if cid is None or pd.isna(cid):
        return None
    cid = str(cid).strip()
    if not cid.isdigit() or len(cid) != 13:
        return None
    return cid

def home_popup_required(page, max_reload=2, timeout=15):
    """
    ✅ พยายามปิด popup
    ✅ ถ้าปิดไม่ได้ → reload หน้าใหม่ แล้วลองใหม่
    ❌ ไม่ lock ตาย
    """
    for attempt in range(1, max_reload + 1):
        start = time.time()
        closed_once = False

        while time.time() - start < timeout:
            try:
                popup_gate(page, soft=True)
                close_popups(page)
                time.sleep(0.4)

                # ✅ ถ้าไม่มี modal ที่ visible → ถือว่าผ่าน
                if not page.locator(".modal:visible").count():
                    log("✅ Home popup cleared / not blocking")
                    return

            except:
                pass

            time.sleep(0.3)

        # ❗ ปิดไม่ได้ → reload
        log(f"♻ Home popup stuck → reload Home (attempt {attempt})")
        page.goto(BASE_URL, wait_until="domcontentloaded")
        page.wait_for_timeout(2000)

    log("⚠ Home popup unresolved → continue (soft)")

def popup_gate(page, timeout=10, soft=True):
    """
    ✅ บังคับจัดการ popup ก่อนทำ action
    - ปิด popup ถ้ามี
    - ถ้า popup หายแล้ว → ไปต่อ
    - ถ้าครบ timeout:
        - soft=True  → log เตือน แล้วไปต่อ
        - soft=False → raise error
    """
    start = time.time()

    while time.time() - start < timeout:
        try:
            warning = page.locator("#warningModal")
            modal = page.locator(".modal")
            backdrop = page.locator(".modal-backdrop")

            # ✅ ถ้ายัง visible จริง → ปิด
            if warning.is_visible() or modal.is_visible():
                close_popups(page)
                time.sleep(0.4)
                continue

            # ✅ backdrop ยังอยู่แต่ไม่ visible → ถือว่าผ่าน
            if backdrop.count() > 0 and not backdrop.first.is_visible():
                return

            # ✅ ไม่มี popup ที่ visible แล้ว
            return

        except Exception:
            # page reload / frame เปลี่ยน
            return

    if soft:
        log("⚠ popup_gate timeout → continue (soft mode)")
        return
    else:
        raise TimeoutError("❌ popup still blocking interaction")


def close_popups(page):
    selectors = [
        "button:has-text('ปิด')",
        "button:has-text('ตกลง')",
        "button:has-text('ยอมรับ')",
        "button:has-text('ยอมรับทั้งหมด')",
        "button:has-text('OK')",
        "button:has-text('Accept')",
        "div:has-text('ปิด')",
        "span:has-text('ปิด')",
        "[role='button']:has-text('ปิด')",
    ]

    for sel in selectors:
        try:
            locs = page.locator(sel)
            for i in range(locs.count()):
                try:
                    locs.nth(i).click(force=True, timeout=800)
                    page.wait_for_timeout(500)
                except:
                    pass
        except:
            pass

def handle_optional_popup(page, timeout=12):
    """
    popup จะโผล่หรือไม่โผล่ก็ได้
    ถ้าโผล่ → ปิด
    ถ้าไม่โผล่ → ไปต่อ (NO ERROR)
    """
    start = time.time()
    while time.time() - start < timeout:
        try:
            if page.locator("#warningModal").is_visible():
                close_popups(page)
                page.wait_for_timeout(300)
            else:
                return
        except:
            return
        page.wait_for_timeout(200)

def wait_popup_must_appear_then_close(page, appear_timeout=12, close_timeout=15):
    """
    popup ต้องโผล่อย่างน้อย 1 ครั้ง
    ถ้าไม่โผล่ => ห้ามทำขั้นตอนต่อ
    """

    popup_selectors = [
        "[role='dialog']",
        "text=เตือน ผู้ใช้งานระบบ",
        "button:has-text('ยอมรับทั้งหมด')",
        "button:has-text('ยอมรับ')",
        "button:has-text('Accept')",
        "button:has-text('ตกลง')",
        "button:has-text('OK')",
        "button:has-text('ปิด')",
        ".modal",
        ".overlay",
    ]

    start = time.time()
    appeared = False

    # ---------- Phase 1: รอ popup โผล่ ----------
    while time.time() - start < appear_timeout:
        for sel in popup_selectors:
            try:
                loc = page.locator(sel).first
                if loc.count() > 0 and loc.is_visible():
                    appeared = True
                    break
            except:
                pass

        if appeared:
            break

        page.wait_for_timeout(200)

    if not appeared:
        raise TimeoutError("❌ Popup did NOT appear (block workflow)")

    log("✔ Popup appeared")

    
    page.wait_for_timeout(300)
    close_popups(page)


    # ---------- Phase 2: ปิด popup ----------
    start = time.time()
    while time.time() - start < close_timeout:
        any_visible = False
        for sel in popup_selectors:
            try:
                loc = page.locator(sel).first
                if loc.count() > 0 and loc.is_visible():
                    any_visible = True
                    close_popups(page)
                    page.wait_for_timeout(300)
                    break
            except:
                pass

        if not any_visible:
            log("✔ Popup closed")
            return

    raise TimeoutError("❌ Popup could not be closed")

def wait_critical_popups(page, timeout=6):
    """
    รอเฉพาะ popup ที่บัง interaction (เร็ว)
    """
    start = time.time()

    selectors = [
        "[role='dialog']",
        "button:has-text('ยอมรับทั้งหมด')",
        "button:has-text('ยอมรับ')",
        "button:has-text('Accept')",
        "button:has-text('ตกลง')",
        "button:has-text('OK')",
        "button:has-text('ปิด')",
    ]

    while time.time() - start < timeout:
        found = False
        for sel in selectors:
            try:
                loc = page.locator(sel).first
                if loc.count() > 0 and loc.is_visible():
                    found = True
                    close_popups(page)
                    page.wait_for_timeout(150)
                    break
            except:
                pass
        if not found:
            return

    log("⚠ critical popup timeout → continue")

def wait_popups_gone(page, timeout=15):
    """
    รอจน dialog / popup / overlay หาย
    ใช้ time.time() แทน page.timeouts (รองรับ Playwright Python)
    """
    start = time.time()

    possible_overlays = [
        "[role='dialog']",
        "button:has-text('ยอมรับทั้งหมด')",
        "button:has-text('ปิด')",
        "div:has-text('เตือน')",
        "div:has-text('ยอมรับ')",
        "div:has-text('Accept')",
        "div:has-text('ไม่ต้องแสดงผลหน้านี้')",
        ".modal",
        ".overlay",
    ]

    while True:
        any_visible = False

        for sel in possible_overlays:
            try:
                loc = page.locator(sel)
                if loc.count() > 0 and loc.first.is_visible():
                    any_visible = True
                    close_popups(page)
                    page.wait_for_timeout(300)
                    break
            except:
                pass

        if not any_visible:
            return

        if time.time() - start > timeout:
            log("⚠ popup still visible but timeout reached → continue")
            return

def wait_table_ready(page, timeout=35000):
    close_popups(page)
    page.wait_for_load_state("domcontentloaded")
    try:
        page.wait_for_load_state("networkidle", timeout=timeout)
    except:
        pass

    selectors = [
        "table thead tr th",
        "table tbody tr",
        "table",
        "[role='table']",
    ]
    last_err = None
    for sel in selectors:
        try:
            page.wait_for_selector(sel, timeout=timeout)
            return
        except Exception as e:
            last_err = e
    if last_err:
        raise last_err
    
    wait_table_ready(page)
    slow_action(0.5)

def click_tab(page, text):
    popup_gate(page, timeout=8, soft=True)
    close_popups(page)
    slow_action(0.3)
    
    tab = page.get_by_text(text, exact=True).first

    tab.wait_for(state="visible", timeout=10000)
    tab.click()
    page.wait_for_load_state("domcontentloaded", timeout=15000)
    page.wait_for_timeout(700)

    try:
        page.wait_for_load_state("domcontentloaded", timeout=10000)
    except:
        pass

    
    popup_gate(page, timeout=8, soft=True)
    close_popups(page)

def merge_if_updated(old_path, new_df):
    new_df = new_df.copy()
    if new_df.empty:
        if os.path.exists(old_path):
            old_df = pd.read_excel(old_path)
            return old_df, False
        return new_df, False

    new_df["item"] = new_df["item"].map(norm_text)
    new_df["amount"] = pd.to_numeric(new_df["amount"], errors="coerce").round(2)

    if not os.path.exists(old_path):
        return new_df, True

    old_df = pd.read_excel(old_path)
    if old_df.empty:
        return new_df, True

    old_df["item"] = old_df["item"].map(norm_text)
    old_df["amount"] = pd.to_numeric(old_df["amount"], errors="coerce").round(2)

    merged = (
        pd.concat([old_df, new_df], ignore_index=True)
        .sort_values(["company_id", "year", "item"])
        .drop_duplicates(["company_id", "year", "item"], keep="last")
        .reset_index(drop=True)
    )

    return merged, not merged.equals(old_df)


def get_years(page):
    ths = page.locator("table thead tr").first.locator("th")
    years = []
    for i in range(ths.count()):
        txt = norm_text(ths.nth(i).text_content())
        if txt.isdigit():
            years.append(int(txt))
    return years


def scrape_table(page, cid, years):
    records = []
    rows = page.locator("table tbody tr")

    for r in range(rows.count()):
        row = rows.nth(r)

        ths = row.locator("th")
        tds = row.locator("td")

        item = ""
        offset = 0

        if ths.count() > 0:
            item = norm_text(ths.first.text_content())
            offset = 0
        else:
            if tds.count() == 0:
                continue
            item = norm_text(tds.first.text_content())
            offset = 1

        if not item:
            continue

        td_count = tds.count()
        if td_count <= offset:
            continue

        data_count = td_count - offset
        stride = 2 if data_count >= len(years) * 2 else 1
        usable = min(len(years), data_count // stride)

        for i in range(usable):
            idx = offset + i * stride
            txt = tds.nth(idx).text_content()
            records.append({
                "company_id": cid,
                "year": years[i],
                "item": item,
                "amount": to_float(txt)
            })

    df = pd.DataFrame(records)
    if not df.empty:
        df["item"] = df["item"].map(norm_text)
        df = df.drop_duplicates(["company_id", "year", "item"])
    return df

def _pick(df, patterns, label="", prefer_exact=True):
    if df is None or df.empty:
        s = pd.Series(dtype="float")
        s.index.name = "year"
        return s

    tmp = df.copy()
    tmp["item"] = tmp["item"].astype(str).map(norm_text)
    tmp["year"] = pd.to_numeric(tmp["year"], errors="coerce")
    tmp["amount"] = pd.to_numeric(tmp["amount"], errors="coerce")

    best = None
    best_score = -1

    for pat in patterns:
        use_regex = pat.startswith("re:")
        expr = pat[3:] if use_regex else re.escape(pat)

        m = tmp[tmp["item"].str.contains(expr, na=False, regex=True)]
        if m.empty:
            continue

        for it in m["item"].unique():
            score = 0
            it_norm = norm_text(it)

            if use_regex:
                score += 2
            if prefer_exact and (it_norm == (pat[3:] if use_regex else pat)):
                score += 10

            for p in patterns:
                p_expr = p[3:] if p.startswith("re:") else p
                if re.search(p_expr if p.startswith("re:") else re.escape(p_expr), it_norm):
                    score += 1

            if score > best_score:
                best_score = score
                best = it

    if best is None:
        if label:
            log(f"⚠️ _pick missing: {label}")
        s = pd.Series(dtype="float")
        s.index.name = "year"
        return s

    s = tmp[tmp["item"] == best].groupby("year")["amount"].last().sort_index()
    s.index.name = "year"
    if label:
        log(f"✔ _pick {label}: '{best}'")
    return s

def _finalize_zdf(df):
    df = df.copy()

    if "บริษัท" not in df.columns:
        df["บริษัท"] = pd.NA
    if "ปี" not in df.columns:
        df["ปี"] = pd.NA

    df["ปี"] = pd.to_numeric(df["ปี"], errors="coerce").astype("Int64")

    for c in MONEY_COLS:
        if c in df.columns:
            df[c] = pd.to_numeric(df[c], errors="coerce").round(2)
        else:
            df[c] = pd.NA

    for c in RATIO_COLS:
        if c in df.columns:
            df[c] = pd.to_numeric(df[c], errors="coerce").round(4)
        else:
            df[c] = pd.NA

    if Z_COL in df.columns:
        df[Z_COL] = pd.to_numeric(df[Z_COL], errors="coerce").round(2)
    else:
        df[Z_COL] = pd.NA

    if "สถานะ" not in df.columns:
        df["สถานะ"] = pd.NA

    for c in COL_ORDER:
        if c not in df.columns:
            df[c] = pd.NA

    df = df[COL_ORDER].sort_values(["บริษัท", "ปี"], kind="stable")
    return df

def compute_zscore(bs, is_df, company_id, company_name):
    if bs is None or bs.empty or is_df is None or is_df.empty:
        return pd.DataFrame(columns=COL_ORDER)

    # =========================
    # Balance Sheet
    # =========================
    CA = _pick(bs, ["สินทรัพย์หมุนเวียน"], "Current Assets")
    CL = _pick(bs, ["หนี้สินหมุนเวียน"], "Current Liab")
    TA = _pick(bs, ["สินทรัพย์รวม"], "Total Assets")
    TL = _pick(bs, ["หนี้สินรวม"], "Total Liab")
    EQ = _pick(bs, ["ส่วนของผู้ถือหุ้น", "ส่วนของเจ้าของ"], "Equity")

    # Retained Earnings
    RE = _pick(bs, [
        "กำไรสะสม",
        r"re:กำไร\s*\(\s*ขาดทุน\s*\)\s*สะสม",
        r"re:กำไรสะสม\s*\(\s*ขาดทุนสะสม\s*\)",
        r"re:กำไรสะสม\s*\(.*สะสม.*\)"
    ], "Retained Earn")

    # =========================
    # Income Statement
    # =========================
    SALES = _pick(is_df, [
        "รายได้รวม",
        "รายได้จากการขาย",
        "รายได้จากการให้บริการ",
        "re:^รายได้(รวม)?$",
        "re:^รายได้จากการขายและการให้บริการ$",
        "re:ยอดขาย|ขาย"
    ], "Sales")

    # --- EBIT (direct pick) ---
    EBIT = _pick(is_df, [
        "EBIT",
        "กำไรจากการดำเนินงาน",
        "กำไรจากกิจกรรมดำเนินงาน",
        r"re:กำไร\s*\(?.*?\)?\s*จากการดำเนินงาน",
    ], "EBIT")

    # ✅ EBIT fallback: PBT + Interest
    if EBIT.isna().all():
        PBT = _pick(is_df, [
            "กำไร(ขาดทุน) ก่อนภาษี",
            r"re:กำไร.*ก่อน.*ภาษี"
        ], "PBT")

        INT_EBIT = _pick(is_df, [
            "ดอกเบี้ยจ่าย",
            "ต้นทุนทางการเงิน",
            "ค่าใช้จ่ายดอกเบี้ย"
        ], "Interest Exp")

        if not PBT.isna().all():
            EBIT = PBT.add(INT_EBIT, fill_value=0)
            log("ℹ EBIT calculated = PBT + Interest")

    # Net Income / Interest / Tax (สำหรับแสดงผล)
    NI = _pick(is_df, [
        r"re:กำไร\s*\(\s*ขาดทุน\s*\)\s*สุทธิ",
        r"re:กำไรสุทธิ\s*\(\s*ขาดทุนสุทธิ\s*\)"
    ], "Net Income")

    INT = _pick(is_df, [
        "ดอกเบี้ยจ่าย",
        "ต้นทุนทางการเงิน",
        "ค่าใช้จ่ายดอกเบี้ย"
    ], "Interest Exp")

    TAX = _pick(is_df, ["ภาษี", "ภาษีเงินได้"], "Tax")

    
    if RE.isna().all() and not NI.isna().all():
        RE = NI.cumsum()
        log("ℹ Retained Earn calculated = cumulative Net Income")


    # =========================
    # Merge
    # =========================
    df = pd.concat(
        [CA, CL, TA, RE, NI, INT, TAX, TL, SALES, EBIT, EQ],
        axis=1
    )

    df.columns = [
        "Current Assets", "Current Liab", "Total Assets",
        "Retained Earn", "Net Income", "Interest Exp", "Tax",
        "Total Liab", "Sales", "EBIT", "Equity"
    ]

    log(f"DEBUG missing counts: {df[['Total Assets','Retained Earn','EBIT']].isna().sum().to_dict()}")

    df.index.name = "year"
    if df.empty or df.dropna(how="all").empty:
        return pd.DataFrame(columns=COL_ORDER)

    # =========================
    # Ratios
    # =========================
    df["บริษัท"] = company_name
    df["Working Cap"] = df["Current Assets"] - df["Current Liab"]

    TA_ = df["Total Assets"].where(df["Total Assets"] != 0)
    TL_ = df["Total Liab"].where(df["Total Liab"] != 0)

    # X1–X5 แยกคำนวณ
    df.loc[TA_.notna(), "X1"] = df.loc[TA_.notna(), "Working Cap"] / TA_
    df.loc[(TA_.notna()) & (df["Retained Earn"].notna()), "X2"] = df["Retained Earn"] / TA_
    df.loc[(TA_.notna()) & (df["EBIT"].notna()), "X3"] = df["EBIT"] / TA_
    df.loc[(TL_.notna()) & (df["Equity"].notna()), "X4"] = df["Equity"] / TL_
    df.loc[(TA_.notna()) & (df["Sales"].notna()), "X5"] = df["Sales"] / TA_

    # =========================
    # Z‑Score (เฉพาะแถวที่ครบ)
    # =========================
    z_valid = (
        df["X1"].notna()
        & df["X2"].notna()
        & df["X3"].notna()
        & df["X4"].notna()
        & df["X5"].notna()
    )

    df.loc[z_valid, Z_COL] = (
        1.2 * df.loc[z_valid, "X1"]
        + 1.4 * df.loc[z_valid, "X2"]
        + 3.3 * df.loc[z_valid, "X3"]
        + 0.6 * df.loc[z_valid, "X4"]
        + 1.0 * df.loc[z_valid, "X5"]
    )

    # =========================
    # Risk label
    # =========================
    def risk_th(z):
        if pd.isna(z):
            return "ไม่ทราบ"
        if z > 2.90:
            return "ปลอดภัย"
        if z >= 1.23:
            return "เฝ้าระวัง"
        return "อันตราย"

    df["สถานะ"] = df[Z_COL].apply(risk_th)

    df = df.reset_index().rename(columns={"year": "ปี"})
    return _finalize_zdf(df)

def _autosize(ws):
    for col in ws.columns:
        max_len = 0
        col_letter = col[0].column_letter
        for cell in col:
            v = "" if cell.value is None else str(cell.value)
            if len(v) > max_len:
                max_len = len(v)
        ws.column_dimensions[col_letter].width = min(max_len + 2, 32)


def _format_numbers(ws):
    headers = [c.value for c in ws[1]]
    idx = {h: i + 1 for i, h in enumerate(headers)}

    for r in range(2, ws.max_row + 1):
        for h in MONEY_COLS:
            if h in idx:
                ws.cell(r, idx[h]).number_format = "#,##0.00"
        for h in RATIO_COLS:
            if h in idx:
                ws.cell(r, idx[h]).number_format = "0.0000"
        if Z_COL in idx:
            ws.cell(r, idx[Z_COL]).number_format = "0.00"


def _colorize(ws):
    headers = [c.value for c in ws[1]]
    if "สถานะ" not in headers:
        return
    status_col = headers.index("สถานะ") + 1

    fills = {
        "ปลอดภัย": PatternFill("solid", fgColor="C6EFCE"),
        "เฝ้าระวัง": PatternFill("solid", fgColor="FFEB9C"),
        "อันตราย": PatternFill("solid", fgColor="FFC7CE"),
        "ไม่ทราบ": PatternFill("solid", fgColor="D9D9D9"),
    }

    for r in range(2, ws.max_row + 1):
        status = ws.cell(r, status_col).value
        fill = fills.get(status)
        if fill:
            for c in range(1, ws.max_column + 1):
                ws.cell(r, c).fill = fill


def _save_new_workbook(df, path):
    wb = Workbook()
    if "Sheet" in wb.sheetnames:
        del wb["Sheet"]
    ws = wb.create_sheet(Z_SCORE_SHEET, 0)

    df = _finalize_zdf(df)
    for r in dataframe_to_rows(df, index=False, header=True):
        ws.append(r)

    ws.freeze_panes = "A2"
    _autosize(ws)
    _format_numbers(ws)
    _colorize(ws)
    wb.save(path)


def build_zscore_pivot(z_df):
    """
    Pivot สำหรับ Dashboard:
    บริษัท | ปี1 | ปี2 | ปี3 ... | เฉลี่ย | แนวโน้ม
    """
    pv = (
        z_df
        .pivot_table(
            index="บริษัท",
            columns="ปี",
            values="Z'-Score",
            aggfunc="last"
        )
        .sort_index()
    )

    
    year_cols = [
        c for c in pv.columns
        if isinstance(c, (int, float, np.integer))
    ]
    year_cols = sorted(year_cols)

    # ===== ค่าเฉลี่ย =====
    pv["เฉลี่ย"] = pv[year_cols].mean(axis=1)

    # ===== แนวโน้ม (Linear Trend / Slope) =====
    trends = []

    for _, row in pv.iterrows():
        y = row[year_cols].values.astype(float)
        x = np.arange(len(y))

        # ต้องมีอย่างน้อย 2 ปี
        if np.count_nonzero(~np.isnan(y)) < 2:
            trends.append("N/A")
            continue

        slope = np.polyfit(x[~np.isnan(y)], y[~np.isnan(y)], 1)[0]

        if slope > 0.05:
            trends.append("📈 ดีขึ้น")
        elif slope < -0.05:
            trends.append("📉 แย่ลง")
        else:
            trends.append("➖ ทรงตัว")

    pv["แนวโน้ม"] = trends

    return pv.reset_index()




def zip_and_base64(zip_path: Path, files: list[Path]):
    if zip_path.exists():
        zip_path.unlink()

    time.sleep(1)

    with zipfile.ZipFile(zip_path, "w", zipfile.ZIP_DEFLATED) as z:
        for f in files:
            if not f.exists():
                raise FileNotFoundError(f"Missing file: {f}")
            if f.stat().st_size <= 0:
                raise RuntimeError(f"File not ready / empty: {f}")
            z.write(f, arcname=f.name)

    # ตรวจว่า ZIP valid จริง
    if not zipfile.is_zipfile(zip_path):
        raise RuntimeError("ZIP corrupted before sending")

    with open(zip_path, "rb") as f:
        return base64.b64encode(f.read()).decode("utf-8")

def build_latest_zscore_table(z_df):
    """
    ตาราง Z'-Score ล่าสุดต่อบริษัท
    """
    latest = (
        z_df
        .sort_values("ปี")
        .dropna(subset=["Z'-Score"])
        .groupby("บริษัท", as_index=False)
        .last()[["บริษัท", "Z'-Score"]]
        .sort_values("Z'-Score")
    )
    return latest

def save_zscore_single_sheet(new_df):
    if new_df is None or new_df.empty:
        log("⚠️ Z-Score empty → skip save")
        return

    # ---------- Load / merge workbook ----------
    if os.path.exists(Z_SCORE_FILE):
        try:
            wb = load_workbook(Z_SCORE_FILE)
        except PermissionError:
            alt = os.path.join(
                OUTPUT_DIR,
                f"zscore_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
            )
            log(f"⚠️ zscore.xlsx locked → save {os.path.basename(alt)}")
            _save_new_workbook(new_df, alt)
            return

        if Z_SCORE_SHEET in wb.sheetnames:
            old_df = pd.read_excel(Z_SCORE_FILE, sheet_name=Z_SCORE_SHEET)
            combined = pd.concat([old_df, new_df], ignore_index=True)
        else:
            combined = new_df.copy()
    else:
        wb = Workbook()
        if "Sheet" in wb.sheetnames:
            del wb["Sheet"]
        combined = new_df.copy()

    # ---------- Finalize + dedup ----------
    combined = _finalize_zdf(combined)
    combined = combined.drop_duplicates(
        ["บริษัท", "ปี"],
        keep="last"
    ).reset_index(drop=True)

    # ---------- Recreate Z_SCORE sheet ----------
    if Z_SCORE_SHEET in wb.sheetnames:
        del wb[Z_SCORE_SHEET]

    ws = wb.create_sheet(Z_SCORE_SHEET, 0)
    for r in dataframe_to_rows(combined, index=False, header=True):
        ws.append(r)

    ws.freeze_panes = "A2"
    _autosize(ws)
    _format_numbers(ws)
    _colorize(ws)

    # ==================================================
    # ✅ IMPORTANT PART (ที่คุณถาม)
    # ==================================================

    # 1) SUMMARY : ค่าเฉลี่ย + แนวโน้มภาพรวม
    add_summary_charts(wb)

    # 2) DASHBOARD : ค่าเฉลี่ย + แนวโน้ม ต่อบริษัท
    pivot_df = build_zscore_pivot(combined)
    add_dashboard_sheet(wb, pivot_df)

    # 3) เปรียบเทียบ Z‑Score ล่าสุด (บริษัทเสี่ยง/ปลอดภัย)
    add_company_zscore_comparison(wb)

    # ==================================================
    # ✅ SAVE WORKBOOK (ต้องอยู่ท้ายสุด)
    # ==================================================
    tmp_path = Z_SCORE_FILE.with_suffix(".xlsx.tmp")
    try:
        wb.save(tmp_path)
        os.replace(tmp_path, Z_SCORE_FILE)
    except PermissionError:
        try:
            if os.path.exists(tmp_path):
                os.remove(tmp_path)
        except:
            pass
        alt = os.path.join(
            OUTPUT_DIR,
            f"zscore_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
        )
        log(f"⚠️ zscore.xlsx locked → save {os.path.basename(alt)}")
        wb.save(alt)

def add_summary_charts(workbook):
    if "Z_SCORE" not in workbook.sheetnames:
        return

    ws_data = workbook["Z_SCORE"]

    # ลบ SUMMARY เดิมถ้ามี
    if "SUMMARY" in workbook.sheetnames:
        del workbook["SUMMARY"]

    ws = workbook.create_sheet("SUMMARY")

    headers = [c.value for c in ws_data[1]]
    idx = {h: i + 1 for i, h in enumerate(headers)}

    status_col = idx.get("สถานะ")
    year_col = idx.get("ปี")
    zcol = idx.get("Z'-Score")
    company_col = idx.get("บริษัท")

    # -------------------------
    # 1) STATUS SUMMARY
    # -------------------------
    ws["A1"] = "สรุปสถานะบริษัท"
    ws.append(["สถานะ", "จำนวน"])

    status_count = {}
    for r in range(2, ws_data.max_row + 1):
        s = ws_data.cell(r, status_col).value
        if s:
            status_count[s] = status_count.get(s, 0) + 1

    start_row = ws.max_row + 1
    for k, v in status_count.items():
        ws.append([k, v])

    pie = PieChart()
    pie.title = "สัดส่วนสถานะบริษัท"
    data = Reference(ws, min_col=2, min_row=start_row, max_row=ws.max_row)
    labels = Reference(ws, min_col=1, min_row=start_row + 1, max_row=ws.max_row)
    pie.add_data(data, titles_from_data=True)
    pie.set_categories(labels)
    ws.add_chart(pie, "D2")

    # -------------------------
    # 2) AVG Z-SCORE BY YEAR
    # -------------------------
    ws["A10"] = "Z-Score เฉลี่ยรายปี"
    ws.append(["ปี", "Z-Score เฉลี่ย"])

    z_by_year = {}
    cnt_by_year = {}

    for r in range(2, ws_data.max_row + 1):
        y = ws_data.cell(r, year_col).value
        z = ws_data.cell(r, zcol).value
        if y and z is not None:
            z_by_year[y] = z_by_year.get(y, 0) + z
            cnt_by_year[y] = cnt_by_year.get(y, 0) + 1

    start_row = ws.max_row + 1
    for y in sorted(z_by_year):
        ws.append([y, round(z_by_year[y] / cnt_by_year[y], 2)])

    line = LineChart()
    line.title = "Z-Score เฉลี่ยรายปี"
    line.y_axis.title = "Z-Score"
    line.x_axis.title = "ปี"

    data = Reference(ws, min_col=2, min_row=start_row, max_row=ws.max_row)
    cats = Reference(ws, min_col=1, min_row=start_row + 1, max_row=ws.max_row)
    line.add_data(data, titles_from_data=True)
    line.set_categories(cats)

    ws.add_chart(line, "D16")

    
    # -------------------------
    # 3) TREND SUMMARY (ง่าย ๆ)
    # -------------------------
    ws["A20"] = "แนวโน้ม Z-Score (ภาพรวม)"
    ws["A20"].font = Font(bold=True)
    ws.append(["ผลลัพธ์", "เงื่อนไข"])

    # ใช้ปีแรกกับปีล่าสุดเท่านั้น
    years = sorted(z_by_year.keys())
    if len(years) >= 2:
        first_year = years[0]
        last_year = years[-1]

        avg_first = z_by_year[first_year] / cnt_by_year[first_year]
        avg_last  = z_by_year[last_year] / cnt_by_year[last_year]

        diff = avg_last - avg_first

        if diff > 0.05:
            trend = "📈 ดีขึ้น"
        elif diff < -0.05:
            trend = "📉 แย่ลง"
        else:
            trend = "➖ ทรงตัว"

        ws.append([trend, f"{first_year} → {last_year}"])
    else:
        ws.append(["N/A", "ข้อมูลไม่พอ"])

def add_company_zscore_comparison(workbook):
    if "Z_SCORE" not in workbook.sheetnames:
        return

    ws_data = workbook["Z_SCORE"]

    if "COMPARE_ZSCORE" in workbook.sheetnames:
        del workbook["COMPARE_ZSCORE"]

    ws = workbook["DASHBOARD"]
    start_row = ws.max_row + 3

    ws[f"A{start_row}"] = "📉 เปรียบเทียบ Z'-Score ล่าสุด (เรียงจากเสี่ยง → ปลอดภัย)"
    ws[f"A{start_row}"].font = Font(bold=True)

    ws.append(["บริษัท", "Z-Score ล่าสุด"])

    headers = [c.value for c in ws_data[1]]
    idx = {h: i + 1 for i, h in enumerate(headers)}

    company_col = idx["บริษัท"]
    year_col = idx["ปี"]
    z_col = idx["Z'-Score"]

    # -----------------------------
    # เลือก Z‑Score ล่าสุดต่อบริษัท
    # -----------------------------
    latest = {}

    for r in range(2, ws_data.max_row + 1):
        cname = ws_data.cell(r, company_col).value
        year = ws_data.cell(r, year_col).value
        z = ws_data.cell(r, z_col).value

        if cname and year and z is not None:
            if cname not in latest or year > latest[cname]["year"]:
                latest[cname] = {"year": year, "z": float(z)}

    # -----------------------------
    # เรียงตาม Z‑Score (ต่ำ → สูง)
    # -----------------------------
    items = sorted(latest.items(), key=lambda x: x[1]["z"])

    ws.append(["บริษัท", "Z-Score"])

    for cname, v in items:
        z = max(min(v["z"], 6), -2)
        ws.append([cname, round(z, 2)])

    # -----------------------------
    # Bar chart แทน line
    # -----------------------------
    chart = BarChart()
    chart.type = "col"
    chart.style = 10
    chart.title = "Z'-Score ล่าสุด (เรียงตามความเสี่ยง)"
    chart.y_axis.title = "Z-Score"
    chart.x_axis.title = "บริษัท"
    chart.width = 26
    chart.height = 14

    data = Reference(
        ws,
        min_col=2,
        min_row=1,
        max_row=ws.max_row
    )
    cats = Reference(
        ws,
        min_col=1,
        min_row=2,
        max_row=ws.max_row
    )

    chart.add_data(data, titles_from_data=True)
    chart.set_categories(cats)

    chart.y_axis.majorGridlines = ChartLines()

    ws.add_chart(chart, "D2")
    
def add_dashboard_sheet(workbook, pivot_df):
    """
    Dashboard แบบอ่านง่าย:
    - ตารางค่าเฉลี่ย + แนวโน้ม ต่อบริษัท
    """
    if pivot_df is None or pivot_df.empty:
        return

    if "DASHBOARD" in workbook.sheetnames:
        del workbook["DASHBOARD"]

    ws = workbook.create_sheet("DASHBOARD")

    ws["A1"] = "📊 Dashboard: ค่าเฉลี่ย & แนวโน้ม Z'-Score"
    ws["A1"].font = Font(size=13, bold=True)

    ws["A3"] = "บริษัท | Z‑Score เฉลี่ย | แนวโน้ม"
    ws["A3"].font = Font(bold=True)

    # ใช้เฉพาะคอลัมน์ที่ต้องการ
    cols = ["บริษัท", "เฉลี่ย", "แนวโน้ม"]
    view_df = pivot_df[cols].copy()

    for r in dataframe_to_rows(view_df, index=False, header=True):
        ws.append(r)

    _autosize(ws)


def add_company_trend_chart(ws, start_row, pivot_df):

    year_cols = [
        i + 1 for i, c in enumerate(pivot_df.columns)
        if isinstance(c, (int, float))
    ]

    if not year_cols:
        return

    chart = LineChart()
    chart.title = "เปรียบเทียบ Z'-Score รายบริษัท"
    chart.y_axis.title = "Z-Score"
    chart.x_axis.title = "ปี"

    data = Reference(
        ws,
        min_col=year_cols[0],
        min_row=start_row + 1,
        max_col=year_cols[-1],
        max_row=start_row + 1   # แถวบริษัทเดียว
    )

    cats = Reference(
        ws,
        min_col=year_cols[0],
        min_row=start_row,
        max_col=year_cols[-1],
        max_row=start_row
    )

    chart.add_data(data, from_rows=True)
    chart.set_categories(cats)

    ws.add_chart(chart, "J9")

def get_done_companies():
    """
    อ่าน zscore.xlsx เพื่อดูว่าบริษัทไหนทำเสร็จแล้ว
    ใช้คอลัมน์ 'บริษัท' เป็น checkpoint
    """
    if not os.path.exists(Z_SCORE_FILE):
        return set()

    try:
        df = pd.read_excel(Z_SCORE_FILE, sheet_name=Z_SCORE_SHEET)
        if "บริษัท" not in df.columns:
            return set()
        done = set(df["บริษัท"].dropna().astype(str).str.strip().unique())
        log(f"📌 Resume mode: found {len(done)} completed companies")
        return done
    except Exception as e:
        log(f"⚠ Cannot read checkpoint zscore.xlsx → {e}")
        return set()

def summarize_run(total_companies):
    done = len(get_done_companies())
    err = len(get_error_companies())

    log("📌 ===== FINAL SUMMARY =====")
    log(f"Total companies : {total_companies}")
    log(f"Completed       : {done}")
    log(f"Errors          : {err}")
    log("===========================")

    return done, err

ERROR_FILE = os.path.join(OUTPUT_DIR, "error_companies.xlsx")
ERROR_BUFFER = []
RUN_ERROR_SET = set()

def log_data_issue(company_name, company_id, category, note=""):
    ERROR_BUFFER.append({
        "บริษัท": company_name,
        "company_id": company_id,
        "category": category,
        "note": note,
        "time": ts()
    })

def restart_browser(p, browser):
    """
    ♻ Restart browser เพื่อล้าง memory leak
    คืนค่า: browser ใหม่, page ใหม่
    """
    try:
        log("♻ Watchdog: restarting browser to prevent memory leak")
        browser.close()
    except:
        pass

    browser = p.chromium.launch(
        headless=HEADLESS,
        args=[
            "--disable-gpu",
            "--no-sandbox",
            "--disable-dev-shm-usage"
        ]
    )
    page = browser.new_page()
    page.set_default_timeout(DEFAULT_TIMEOUT)
    page.goto(BASE_URL, wait_until="domcontentloaded")
    page.wait_for_load_state("networkidle", timeout=15000)
    page.wait_for_timeout(800)

    return browser, page

# =========================
# ERROR CLASSIFICATION
# =========================

# ❌ error ที่ไม่มีทางสำเร็จ → ข้ามถาวร
FATAL_ERRORS = {
    "SEARCH_NOT_FOUND",
    "NO_BALANCE_SHEET",
    "NO_INCOME_STATEMENT",
    "VALID_NO_FINANCIAL",
}

# ✅ error ที่ retry ได้
RETRYABLE_ERRORS = {
    "COMPANY_TIMEOUT",
    "PAGE_TIMEOUT",
    "PLAYWRIGHT_TIMEOUT",
    "TARGET_CLOSED",
    "UNKNOWN_ERROR",
}

def get_error_companies():
    if not os.path.exists(ERROR_FILE):
        return set()

    try:
        df = pd.read_excel(ERROR_FILE)
        if "บริษัท" not in df.columns:
            return set()
        errs = set(df["บริษัท"].dropna().astype(str).str.strip().unique())
        log(f"📌 Resume mode: found {len(errs)} error companies")
        return errs
    except Exception as e:
        log(f"⚠ Cannot read error checkpoint → {e}")
        return set()
    
def get_fatal_error_companies():
    """
    อ่าน error_companies.xlsx
    คืนเฉพาะบริษัทที่เป็น fatal error (ไม่ต้องทำอีก)
    """
    if not os.path.exists(ERROR_FILE):
        return set()

    try:
        df = pd.read_excel(ERROR_FILE)
        if df.empty or "บริษัท" not in df.columns or "error" not in df.columns:
            return set()

        fatal_df = df[df["error"].isin(FATAL_ERRORS)]
        return set(
            fatal_df["บริษัท"]
            .dropna()
            .astype(str)
            .str.strip()
            .unique()
        )
    except Exception as e:
        log(f"⚠ Cannot read fatal error list → {e}")
        return set() 

def save_error_company(company_name, company_id, reason):
    ERROR_BUFFER.append({
        "บริษัท": company_name,
        "company_id": company_id,
        "error": reason,
        "time": ts()
    })

    RUN_ERROR_SET.add(company_name)

def flush_error_buffer():
    if not ERROR_BUFFER:
        return

    if os.path.exists(ERROR_FILE):
        df = pd.read_excel(ERROR_FILE)
        df = pd.concat([df, pd.DataFrame(ERROR_BUFFER)], ignore_index=True)
    else:
        df = pd.DataFrame(ERROR_BUFFER)

    df = df.drop_duplicates(["บริษัท"], keep="last")
    safe_to_excel(df, ERROR_FILE)


def read_error_summary(error_file):
    error_file = Path(error_file)
    if not error_file.exists():
        return 0, {}

    df = pd.read_excel(error_file)
    if df.empty:
        return 0, {}

    total_error = len(df)
    by_reason = (
        df["error"].value_counts().to_dict()
        if "error" in df.columns
        else {}
    )

    return total_error, by_reason

POWER_AUTOMATE_WEBHOOK = "https://default5db8bf0e85924ed082b2a6d4d77933.d4.environment.api.powerplatform.com:443/powerautomate/automations/direct/workflows/d1cb9b9204a24ba88c68df056d6ba648/triggers/manual/paths/invoke?api-version=1&sp=%2Ftriggers%2Fmanual%2Frun&sv=1.0&sig=Zace1Nf1nOM0XjWuwQPrlvY_9a4wJSXHH0w8tNo41Hg"  

def build_mail_body(
    status,
    total,
    completed,
    errors,
    start_time,
    success_rate,
    now_time,
    eta,
    output_path
):
    success_rate = round((completed / total) * 100, 2) if total else 0

    return f"""Status: {status}

Summary
-----------------------
Total companies : {total}
Completed       : {completed}
Errors          : {errors}
Success rate    : {success_rate} %

Timing
-----------------------
Start : {start_time}
Now   : {now_time}

Estimated Finish
-----------------------
ETA   : {eta}

Output
-----------------------
{output_path}
"""

def notify_outlook_batch_report(
    *,
    status,                     # COMPLETED | PAUSED_TIME_WINDOW
    total,
    done,
    errors,
    remaining,
    start_dt,
    end_dt,
    elapsed_time_str,
    avg_time_per_company_sec,
    eta_str,
):
    attachments = build_attachments()

    payload = {
        "subject": (
            "[END OF DAY] Z-Score Batch Paused (Time Window Reached)"
            if status == "PAUSED_TIME_WINDOW"
            else "[PROD] Z-Score Batch Completed"
        ),
        "status": status,

        "progress": {
            "total": total,
            "completed": done,
            "errors": errors,
            "remaining": remaining,
            "progress_pct": round((done / total) * 100, 2) if total else 0,
        },

        "program_timing": {
            "start": start_dt.strftime("%Y-%m-%d %H:%M:%S"),
            "end": end_dt.strftime("%Y-%m-%d %H:%M:%S"),
            "elapsed": elapsed_time_str,
        },

        "performance": {
            "avg_time_per_company_sec": (
                round(avg_time_per_company_sec, 2)
                if avg_time_per_company_sec else None
            )
        },

        "eta": eta_str,
        "attachments": attachments,
    }

    try:
        requests.post(POWER_AUTOMATE_WEBHOOK, json=payload, timeout=30)
        log("📧 Outlook PROD batch report sent (with attachments)")
    except Exception as e:
        log(f"❌ Failed to send Outlook batch report: {e}")


def calc_runtime_info(start_dt, now_dt, completed, total):
    elapsed = now_dt - start_dt

    days = elapsed.days
    hours, rem = divmod(elapsed.seconds, 3600)
    minutes = rem // 60

    elapsed_str = f"{days} days {hours} hours {minutes} minutes"

    # ETA แบบง่าย (ถ้ามี progress)
    if completed > 0:
        avg_sec_per_company = elapsed.total_seconds() / completed
        remaining = max(total - completed, 0)
        eta_dt = now_dt + timedelta(seconds=avg_sec_per_company * remaining)
        eta_str = eta_dt.strftime("%Y-%m-%d %H:%M:%S")
    else:
        eta_str = "-"

    return {
        "now": now_dt.strftime("%Y-%m-%d %H:%M:%S"),
        "elapsed_str": elapsed_str,
        "runtime_days": days,
        "eta": eta_str
    }

def ensure_datetime(t):
    if isinstance(t, str):
        return datetime.fromisoformat(t)
    return t

def notify_outlook_crash(error, start_time):
    attachments = build_attachments()

    payload = {
        "subject": "[CRASH] Z-Score Batch Job Failed",
        "status": "CRASHED",
        "error": str(error),
        "error_type": type(error).__name__,
        "start_time": (
            start_time.strftime("%Y-%m-%d %H:%M:%S")
            if hasattr(start_time, "strftime")
            else str(start_time)
        ),
        "crash_time": ts(),
        "attachments": attachments,
    }

    try:
        requests.post(
            POWER_AUTOMATE_WEBHOOK,
            json=payload,
            timeout=20
        )
        log("🚨 CRASH notification sent to Outlook (with attachments)")
    except Exception as e:
        log(f"❌ Failed to send crash notify: {e}")

def file_to_base64(path: Path):
    """
    แปลงไฟล์ -> Base64 สำหรับแนบใน Outlook (Power Automate)
    """
    try:
        if not path.exists():
            return None
        with open(path, "rb") as f:
            return base64.b64encode(f.read()).decode("utf-8")
    except Exception as e:
        log(f"⚠ Cannot encode file {path}: {e}")
        return None
    

def build_attachments():
    zip_path = OUTPUT_DIR / "zscore_reports.zip"

    files = []
    if Z_SCORE_FILE.exists():
        files.append(Z_SCORE_FILE)

    error_file = Path(ERROR_FILE)
    if error_file.exists():
        files.append(error_file)

    if not files:
        return []

    zip_b64 = zip_and_base64(zip_path, files)

    return [{
        "name": "zscore_reports.zip",
        "contentBytes": zip_b64
    }]

def notify_outlook_power_automate(
    *,
    subject: str,
    status: str,
    payload_extra: dict | None = None
):
    """
    ✅ ส่ง Outlook ผ่าน Power Automate (HTTP Webhook)
    """

    # ✅ แนบไฟล์เฉพาะ COMPLETED และ TEST
    if status in ["COMPLETED", "COMPLETED_WITH_ERRORS", "TEST", "PAUSED_TIME_WINDOW"]:
        attachments = build_attachments()
    else:
        attachments = []

    payload = {
        "subject": subject,
        "status": status,
        "timestamp": ts(),
        "attachments": attachments,
        "has_attachments": bool(attachments),
    }

    if payload_extra:
        payload.update(payload_extra)

    try:
        resp = requests.post(
            POWER_AUTOMATE_WEBHOOK,
            json=payload,
            timeout=30
        )
        log(f"📨 Outlook response: {resp.status_code}")
        if resp.text:
            log(f"📨 Outlook response body: {resp.text}")

    except Exception as e:
        log(f"❌ Power Automate notify failed: {e}")

def save_zscore_per_company(zdf, cid):
    """
    Export Z-Score รายบริษัท เป็นไฟล์ Excel แยก
    """
    if zdf is None or zdf.empty:
        log(f"⚠️ Z-Score empty for {cid} → skip company file")
        return

    out_path = os.path.join(OUTPUT_DIR, cid, f"zscore_{cid}.xlsx")

    wb = Workbook()
    ws = wb.active
    ws.title = "Z_SCORE"

    zdf = _finalize_zdf(zdf)

    for r in dataframe_to_rows(zdf, index=False, header=True):
        ws.append(r)

    ws.freeze_panes = "A2"
    _autosize(ws)
    _format_numbers(ws)
    _colorize(ws)

    wb.save(out_path)
    log(f"💾 Saved company Z-Score → {out_path}")

ids_df = pd.read_excel(
    INPUT_IDS_FILE,
    sheet_name=IDS_SHEET,
    dtype={ID_COL: str}
)

ids_df[ID_COL] = ids_df[ID_COL].astype(str).str.strip().str.zfill(13)
ids_df["CompanyName"] = ids_df["CompanyName"].astype(str).str.strip()

company_list = ids_df[[ID_COL, "CompanyName"]].to_dict("records")

filtered_companies = []

for rec in company_list:
    cname = rec.get("CompanyName", "").strip()

    cid_raw = rec.get(ID_COL)
    if cid_raw is None or pd.isna(cid_raw):
        log_data_issue(cname, cid_raw, "INVALID_NO_ID", "No company_id from source")
        continue

    cid = str(cid_raw).strip()
    if not cid.isdigit() or len(cid) != 13:
        log_data_issue(cname, cid, "INVALID_NO_ID", "Company_id not 13 digits")
        continue

    if looks_like_person_name(cname):
        log_data_issue(cname, cid, "INVALID_PERSON_NAME", "Personal name")
        continue

    if looks_like_university(cname):
        log_data_issue(cname, cid, "INVALID_UNIVERSITY", "University / Institute")
        continue

    filtered_companies.append({
        ID_COL: cid,
        "CompanyName": cname
    })

company_list = filtered_companies
log(f"✅ After pre-filter: {len(company_list)} companies remain")

# =====================================================
# DEDUPLICATE BY COMPANY_ID (ก่อนเปิด Browser)
# =====================================================
dedup_map = {}

for rec in company_list:
    cid = rec[ID_COL]
    cname = rec["CompanyName"]

    if cid not in dedup_map:
        dedup_map[cid] = rec
    else:
        old_name = dedup_map[cid]["CompanyName"]

        if len(cname) > len(old_name):
            dedup_map[cid] = rec

        log_data_issue(
            cname,
            cid,
            "DUPLICATE_ID",
            "Duplicate company_id, merged"
        )

company_list = list(dedup_map.values())
log(f"✅ After ID dedup: {len(company_list)} unique companies remain")

done_companies = get_done_companies()
fatal_error_companies = get_fatal_error_companies()

if TEST_MODE:
    company_list = company_list[:TEST_LIMIT]
    log(f"🧪 TEST MODE → {len(company_list)} companies")

def ensure_company_page(page, cid, timeout=30):
    """
    ทำให้แน่ใจว่าเข้า 'หน้าบริษัท' แล้วจริง
    ถ้ายัง → กด Enter ซ้ำแบบปลอดภัย
    """
    start = time.time()

    while time.time() - start < timeout:
        try:
            # ✅ menu2 มีเฉพาะหน้าบริษัท
            if page.locator("#menu2").count() > 0:
                return True
        except:
            pass

        search = page.locator("form.form-group.search.lg input.form-control").first
        page.evaluate("(el) => el.focus()", search)
        page.keyboard.press("Enter")
        page.wait_for_timeout(800)

    raise TimeoutError(f"Cannot enter company page for {cid}")

def type_search_with_retry(page, browser, cid, retries=2):
    for attempt in range(1, retries + 1):
        log(f"⌨ Search attempt {attempt}/{retries} for {cid}")

        page = ensure_page_alive(page, browser)

        # ✅ รีเซ็ต Home ก่อนทุกครั้ง
        page.goto(BASE_URL, wait_until="domcontentloaded")
        page.wait_for_timeout(1500)

        search = ensure_home_search_ready(page, browser, timeout=180)

        try:
            page.evaluate("(el) => el.focus()", search)
        except:
            pass

        search.fill("")
        page.wait_for_timeout(300)

        slow_type(search, cid)
        page.keyboard.press("Enter")
        slow_action(1.2)

        popup_gate(page, soft=True)

        if page.locator("#menu2").count() > 0:
            log("✅ Entered company page")
            return True

        log("🔁 Not in company page yet")

    raise TimeoutError(f"❌ Cannot search company {cid}")

def clear_error_company(company_name):
    if not os.path.exists(ERROR_FILE):
        return
    df = pd.read_excel(ERROR_FILE)
    df = df[df["บริษัท"] != company_name]
    safe_to_excel(df, ERROR_FILE)

# =========================
# SAFE EXCEL WRITE (WINDOWS)
# =========================
def safe_to_excel(df, path, retries=5, sleep=0.5):
    for i in range(retries):
        try:
            df.to_excel(path, index=False, engine="openpyxl")
            return True
        except PermissionError:
            if i == retries - 1:
                raise
            time.sleep(sleep)

def is_within_run_window():
    now = datetime.now().time()
    return RUN_START_TIME <= now <= RUN_END_TIME


def disable_all_scroll(page):
    page.evaluate("""
        () => {
            // ปิด scroll ทุกชนิด
            window.scrollTo = () => {};
            window.scrollBy = () => {};
            Element.prototype.scrollIntoView = () => {};
            document.documentElement.style.scrollBehavior = "auto";
            document.body.style.scrollBehavior = "auto";
        }
    """)

# =====================================================
# MAIN
# =====================================================

run_start_dt = datetime.now()

try:
    with sync_playwright() as p:
        log("🚀 Launch browser")
        browser = p.chromium.launch(
            headless=HEADLESS,
            args=[
                "--disable-gpu",
                "--no-sandbox",
                "--disable-dev-shm-usage"
            ]
        )
        page = browser.new_page()
        page.set_default_timeout(DEFAULT_TIMEOUT)

        page.goto(BASE_URL, wait_until="domcontentloaded")
        page.wait_for_timeout(1200)

        disable_all_scroll(page)

        stopped_by_time_window = False

        # =========================
        # MAIN COMPANY LOOP
        # =========================
        for i, rec in enumerate(company_list, start=1):
            # ---- TIME WINDOW CHECK ----
            if not is_within_run_window():
                log("⏰ Outside run window → graceful shutdown")
                stopped_by_time_window = True
                break

            # ---- WATCHDOG ----
            if i % WATCHDOG_RESTART_EVERY == 0:
                browser, page = restart_browser(p, browser)

            log(f"❤️ Progress {i}/{len(company_list)} companies")

            cid = rec[ID_COL]
            cname = rec["CompanyName"]
            company_start_time = time.time()
            # ---- CHECKPOINT ----
            if cname in done_companies:
                log(f"⏭ Skip (already done): {cname}")
                continue

            if cname in fatal_error_companies:
                log(f"⛔ Skip (known error): {cname}")
                continue

            log(f"🔎 Start company_id = {cid} | {cname}")
            success = False

            # =========================
            # PER‑COMPANY RETRY LOOP
            # =========================
            for attempt in range(1, MAX_RETRY + 1):

                if time.time() - company_start_time > COMPANY_HARD_TIMEOUT:
                    log("⏱ Company hard-timeout exceeded → skip")
                    save_error_company(cname, cid, "COMPANY_TIMEOUT")
                    success = True
                    break

                log(f"🔁 Attempt {attempt}/{MAX_RETRY}")

                try:
                    # ---------- STEP 1: HOME / SEARCH ----------
                    home_popup_required(page)
                    popup_gate(page, soft=True)

                    type_search_with_retry(page, browser, cid, retries=2)

                    # ---------- STEP 2: OPEN MENU ----------
                    page.wait_for_selector("#menu2", timeout=25000)
                    slow_click(page.locator("#menu2"))
                    page.wait_for_load_state("domcontentloaded", timeout=20000)
                    page.wait_for_timeout(800) 

                    if SAFE_MODE:
                        handle_optional_popup(page, timeout=8)

                    # ---------- STEP 3: BALANCE SHEET ----------
                    click_tab(page, "งบแสดงฐานะการเงิน")
                    page.wait_for_selector("table", timeout=20000)
                    page.wait_for_timeout(800)

                    bs_years = get_years(page)
                    if not bs_years:
                        save_error_company(cname, cid, "NO_BALANCE_SHEET")
                        flush_error_buffer()
                        success = True
                        break

                    bs = scrape_table(page, cid, bs_years)
                    log(f"📊 BS rows={len(bs)} years={bs_years}")

                    if bs.empty:
                        save_error_company(cname, cid, "VALID_NO_FINANCIAL")
                        flush_error_buffer()
                        success = True
                        break

                    # ---------- STEP 4: INCOME STATEMENT ----------
                    click_tab(page, "งบกำไรขาดทุน")
                    page.wait_for_selector("table", timeout=20000)
                    page.wait_for_timeout(800)

                    is_years = get_years(page)
                    if not is_years:
                        save_error_company(cname, cid, "NO_INCOME_STATEMENT")
                        flush_error_buffer()
                        success = True

                    is_df = scrape_table(page, cid, is_years)
                    log(f"📊 IS rows={len(is_df)} years={is_years}")

                    if is_df.empty:
                        save_error_company(cname, cid, "VALID_NO_FINANCIAL")
                        flush_error_buffer()
                        success = True
                        break

                    # ---------- STEP 5: SAVE RAW ----------
                    cid_dir = OUTPUT_DIR / cid
                    cid_dir.mkdir(exist_ok=True)

                    bs_final, bs_up = merge_if_updated(cid_dir / f"balance_sheet_{cid}.xlsx", bs)
                    if bs_up:
                        safe_to_excel(bs_final, cid_dir / f"balance_sheet_{cid}.xlsx")

                    is_final, is_up = merge_if_updated(cid_dir / f"income_statement_{cid}.xlsx", is_df)
                    if is_up:
                        safe_to_excel(is_final, cid_dir / f"income_statement_{cid}.xlsx")

                    # ---------- STEP 6: Z‑SCORE ----------
                    zdf = compute_zscore(bs_final, is_final, cid, cname)
                    save_zscore_single_sheet(zdf)
                    save_zscore_per_company(zdf, cid)

                    log(f"✅ Done company_id = {cid}")
                    clear_error_company(cname)
                    success = True
                    break

                except Exception as e:
                    err_msg = str(e)
                    err_type = type(e).__name__

                    log(f"❌ Attempt {attempt} Error: {err_type}: {err_msg}")
                    traceback.print_exc()

                    # ===== CASE 1: ค้นหาไม่เจอ → ข้ามถาวร =====
                    if isinstance(e, TimeoutError) and "Cannot search company" in err_msg:
                        save_error_company(cname, cid, "SEARCH_NOT_FOUND")
                        flush_error_buffer()
                        success = True   # ✅ ถือว่าจบบริษัทนี้แล้ว
                        break

                    # ===== CASE 2: หน้าเข้าได้ แต่ไม่มีงบ =====
                    if "NO_BALANCE_SHEET" in err_msg or "NO_INCOME_STATEMENT" in err_msg:
                        save_error_company(cname, cid, "VALID_NO_FINANCIAL")
                        flush_error_buffer()
                        success = True
                        break
                    # ===== CASE 3: timeout / page error → retry ได้ =====
                    if isinstance(e, TimeoutError) or "TargetClosedError" in err_msg:
                        log("🔁 Retryable timeout/page error → retry search")
                        page = ensure_page_alive(page, browser)
                        continue

                    # ===== CASE 4: อื่น ๆ =====
                    if attempt == MAX_RETRY:
                        save_error_company(cname, cid, "MAX_RETRY_EXCEEDED")
                        flush_error_buffer()

                    # ✅ only page‑level recovery here
                    page = ensure_page_alive(page, browser)

            if not success:
                save_error_company(cname, cid, "MAX_RETRY_EXCEEDED")
                flush_error_buffer()

            # ---- RETURN HOME FOR NEXT COMPANY ----
            page.goto(BASE_URL, wait_until="domcontentloaded")
            page.wait_for_timeout(1200)

        # =====================================================
        # 🔴 POST‑BATCH: MUST BE OUTSIDE LOOP
        # =====================================================
        flush_error_buffer()

        done = len(get_done_companies())
        run_end_dt = datetime.now()

        # ===== POST-RUN STATS =====
        errors = len(RUN_ERROR_SET)
        processed = done + errors
        remaining = len(company_list) - processed

        now_dt = run_end_dt

        # ⏱ ใช้ไปแล้วกี่วัน กี่ชั่วโมง
        elapsed_str, _ = format_elapsed(run_start_dt, now_dt)

        # ⚡ เวลาเฉลี่ยต่อบริษัท
        avg_sec = estimate_avg_time_per_company(
            start_dt=run_start_dt,
            now_dt=now_dt,
            processed=processed
        )

        # 📅 ETA
        eta_str = "N/A"
        if avg_sec and remaining > 0:
            eta_dt = estimate_eta(now_dt, avg_sec * remaining)
            eta_str = eta_dt.strftime("%Y-%m-%d %H:%M")

        # ===== FINAL OUTLOOK NOTIFICATION =====
        subject = (
            "[END OF DAY] Z-Score Batch Paused (Time Window Reached)"
            if stopped_by_time_window
            else "[PROD] Z-Score Batch Completed"
        )

        status = (
            "PAUSED_TIME_WINDOW"
            if stopped_by_time_window
            else ("COMPLETED_WITH_ERRORS" if errors > 0 else "COMPLETED")
        )

        notify_outlook_power_automate(
            subject=subject,
            status=status if not TEST_MODE else "TEST",
            
            payload_extra={
                "summary": {
                    "total": len(company_list),
                    "completed": done,
                    "errors": errors,
                    "remaining": remaining,
                    "success_rate": round((done / len(company_list)) * 100, 2)
                    if len(company_list) else 0
                },
                "timing": {
                    "start": run_start_dt.strftime("%Y-%m-%d %H:%M:%S"),
                    "end": run_end_dt.strftime("%Y-%m-%d %H:%M:%S"),
                    "elapsed": elapsed_str,
                },
                "performance": {
                    "avg_time_per_company_sec": round(avg_sec, 2) if avg_sec else None
                },
                "eta": eta_str,
                "mode": "TEST" if TEST_MODE else "PROD",
                "output_path": str(OUTPUT_DIR)
            }
        )

        summarize_run(len(company_list))
        log("🛑 Close browser")
        browser.close()


except Exception as fatal:
    log("🔥 PROGRAM CRASHED (FATAL)")
    traceback.print_exc()

    try:
        flush_error_buffer()
    except:
        pass

    notify_outlook_power_automate(
        subject="[CRASH] Z-Score Batch Job Failed",
        status="CRASHED",
        payload_extra={
            "error": str(fatal),
            "error_type": type(fatal).__name__,
            "start_time": (
                run_start_dt.strftime("%Y-%m-%d %H:%M:%S")
                if hasattr(run_start_dt, "strftime")
                else str(run_start_dt)
            ),
            "crash_time": ts(),
        }
    )

    raise